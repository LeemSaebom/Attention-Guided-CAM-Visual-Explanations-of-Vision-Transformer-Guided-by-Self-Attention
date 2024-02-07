import h5py
import torch
import numpy as np
import os
import tqdm
import gc
import timm
import argparse
import random
import torchvision.transforms as transforms
import openpyxl
import torch.utils.model_zoo as model_zoo

# Before calculating the ABPC score, the h5 file for the target method is required.
parser = argparse.ArgumentParser(description='perturbation')
parser.add_argument('--method', type=str, choices=['agcam', 'lrp', 'rollout']) 
parser.add_argument('--h5_root', type=str, required=True)
parser.add_argument('--csv', type=str, default="True", choices=['True', 'False'])
parser.add_argument('--file', type=str, default="True", choices=['True', 'False'])
args = parser.parse_args()


MODEL = 'vit_base_patch16_224'
DEVICE = 'cuda'
device = 'cuda' if torch.cuda.is_available() else 'cpu'

# Set seed
def set_seed(seed):
    torch.backends.cudnn.deterministic = True
    torch.backends.cudnn.benchmark = False
    torch.manual_seed(seed)
    torch.cuda.manual_seed_all(seed)
    np.random.seed(seed)
    random.seed(seed)
set_seed(777)
if device == DEVICE:
    gc.collect()
    torch.cuda.empty_cache()
print("device: " +  device)


root = args.h5_root
file_name = ""

# Information for ILSVRC
mean = [0.5, 0.5, 0.5]
std = [0.5, 0.5, 0.5]
normalize = transforms.Compose([
    transforms.Normalize(mean=mean, std=std)
])

state_dict = model_zoo.load_url('https://github.com/rwightman/pytorch-image-models/releases/download/v0.1-vitjx/jx_vit_base_p16_224-80ecf9dd.pth', progress=False, map_location=device)

class_num=1000
file_name +="ILSVRC"


if args.method=="agcam":
    file_name +="_agcam"
elif args.method=="lrp":
    file_name+="_lrp"
elif args.method=="rollout":
    file_name+='_rollout'

# Load the ViT model to calculate the probability score
model = timm.create_model(MODEL, pretrained=True, num_classes=class_num).to(device)
model.load_state_dict(state_dict, strict=True)
model.eval()

# The path for h5 file
h5_filepath = os.path.join(root, file_name+".hdf5")

print("Read h5 file from: ", h5_filepath)
file = h5py.File(h5_filepath)

names = list(file['image'].keys())
print("The size of dataset is: ", len(names))

# The perturbation scope is 0.0 to 1.0 with an interval of 0.1
step = 0.1
scope = [0.0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0]

# ABPC is defined by the AUC of MoRF subtracted from the AUC of LeRF
# Matrices to save the probability score of each image with each perturbation step
LeRF = np.zeros((len(scope), len(names)))
MoRF = np.zeros((len(scope), len(names)))

base=224*224

for idx, name in enumerate(tqdm.tqdm(names)):
    data = torch.tensor(np.array(file['/image'][name])).to(device)    # The original input image
    vis = torch.tensor(np.array(file['/cam'][name])).to(device)       # The heatmap visualization
    target = torch.tensor(np.array(file['/label'][name])).to(device)  # The label class
    c, h, w = vis.shape
    data_norm = normalize(data).to(device)
    for i in range(len(scope)):                             # Perturbation step: [0.0, 0.1, 0.2, 0.3, 0.4, 0.5, 0.6, 0.7, 0.8, 0.9, 1.0]
        k = int(base*scope[i])
        mask = vis.view(c, -1).clone() + 1.0
        value, indices = torch.topk(mask, k)                # Select top k pixels in the heatmap
        mask[0, indices]=0
        mask = torch.where(mask==0.0, 0.0, 1.0)             # Mask the image
        mask = mask.reshape([h, w])
        masked_data = data_norm * mask
        pred = model(masked_data)
        prob = torch.softmax(pred, dim=1)
        MoRF[i, idx] = prob[:, target.item()].item()        # Probability score of the masked image

        mask = vis.view(c, -1).clone() * -1.0 - 1.0
        value, indices = torch.topk(mask,k)                 # Select bottom k pixels in the heatmap
        mask[0, indices]=0
        mask = torch.where(mask==0.0, 0.0, 1.0)             # Mask the Image
        mask = mask.reshape([h,w])
        masked_data = data_norm * mask

        pred = model(masked_data)
        prob = torch.softmax(pred, dim=1)
        LeRF[i, idx] = prob[:, target.item()].item()        # Probability score of the masked image

# Print the results
LeRF_mean = np.mean(LeRF, axis=1)
MoRF_mean = np.mean(MoRF, axis=1)
auc_LeRF = np.sum(LeRF_mean) / (len(scope))
auc_MoRF = np.sum(MoRF_mean) / (len(scope))
print("AUC of LeRF per step: ")
print(LeRF_mean)
print("AUC of LeRF: ")
print(auc_LeRF)
print("AUC of MoRF per step: ")
print(MoRF_mean)
print("AUC of MoRF: ")
print(auc_MoRF)

abpc = np.subtract(LeRF, MoRF)
abpc_mean = np.mean(abpc, axis=1)
abpc_score = np.sum(abpc_mean)/(len(scope))
print("ABPC per step: ")
print(abpc_mean)
print("ABPC score: ")
print(abpc_score)

# Save the result in a text file.
if args.file=="True":
    file = open(os.path.join(root, file_name+'_ABPC.txt'), 'a')
    file.write("\n\n=======================================")
    file.write("\n")
    file.write("\npixel perturbation")
    file.write("\nscope: ")
    for context in scope:
        file.write(str(context) + " ")
    file.write("\nAIC of LeRF per step: ")
    for context in LeRF_mean:
        file.write(str(context) + " ")
    file.write("\nAUC of LeRF: ")
    file.write(str(auc_LeRF))
    file.write("\nAUC of MoRF per step: ")
    for context in MoRF_mean:
        file.write(str(context) + " ")
    file.write("\nAUC of MoRF: ")
    file.write(str(auc_MoRF))
    file.write("\nABPC per step: ")
    for context in abpc_mean:
        file.write(str(context) + " ")
    file.write("\nABPC score: ")
    file.write(str(abpc_score))


# If you want to see the ABPC score of each of the images in the dataset, please create csv file.
if args.csv == "True":
    csv_save = os.path.join(root, file_name + "_ABPC.xlsx")
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = str(step)
    sheet.cell(row=1, column=1).value = "file name"
    for i in range(len(scope)):
        sheet.cell(row=1, column=i+2).value = str(scope[i])
    sheet.cell(row=1, column=len(scope)+3).value = "ABPC"
    sum_ = np.sum(abpc, axis=0) / (len(scope)+1)
    print("\nsaving the csv.....")
    for idx, name in enumerate(tqdm.tqdm(names)):
        sheet.cell(row=idx+2, column=1).value = name
        for i in range(len(scope)):
            sheet.cell(row=idx+2, column=i+2).value = abpc[i, idx]
        sheet.cell(row=idx+2, column=len(scope)+3).value = sum_[idx]
    wb.save(csv_save)



